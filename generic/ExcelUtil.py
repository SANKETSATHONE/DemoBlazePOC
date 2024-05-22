import openpyxl
import os
from generic.fileFunctionUtil import FileFunctionUtil


class ExcelUtil:
    @staticmethod
    def get_excel_data(excelFileName, testCaseID):
        TCRowNum = 1
        TCID = dict()
        desired_path = os.path.join(os.path.dirname(FileFunctionUtil.get_dynamic_path("DemoBlazePOC") + "\\testData"),
                                    'testData',
                                    'excel')
        workbook = openpyxl.load_workbook(desired_path + '\\' + excelFileName)
        sheet = workbook.active
        while True:
            TCValue = sheet.cell(row=TCRowNum + 1, column=1)
            if TCValue.value == None:
                break
            if TCRowNum > 298:
                break
            TCID[TCValue.value] = TCRowNum + 1
            TCRowNum += 2
        TestData = dict()
        TCIDRowNum = TCID[testCaseID]
        ColumnNum = 1
        while True:
            Key = sheet.cell(row=TCIDRowNum - 1, column=ColumnNum)
            Value = sheet.cell(row=TCIDRowNum, column=ColumnNum)
            if Key.value == None:
                break
            if ColumnNum > 33:
                break
            TestData[Key.value] = Value.value
            ColumnNum += 1
        return TestData
